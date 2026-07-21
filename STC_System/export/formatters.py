"""
export/formatters.py
────────────────────
Low-level openpyxl formatting helpers used by all sheet writers.

Provides:
  • Header / data / error cell formats
  • RTL sheet direction
  • Auto-fit column widths
  • Frozen header row
  • Filters on all columns
  • Conditional formatting (colour scales, data bars, highlight)
  • Chart factory (bar, pie, line)
"""
from __future__ import annotations

from typing import List, Optional

import openpyxl
from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint
from openpyxl.formatting.rule import (
    ColorScaleRule,
    DataBarRule,
    CellIsRule,
    FormulaRule,
)
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    GradientFill,
    PatternFill,
    Side,
    numbers,
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

import polars as pl

# ─── Colour palette ──────────────────────────────────────────────────────────

NAVY       = "1f2d5a"  # Header background
WHITE      = "FFFFFF"
LIGHT_BLUE = "dce6f1"  # Alternating row A
LIGHT_GREY = "f5f5f5"  # Alternating row B
RED        = "da3633"
GREEN      = "238636"
AMBER      = "d29922"
PURPLE     = "8957e5"
TEAL       = "1abc9c"
DARK_BG    = "0d1117"

# ─── Cell style helpers ───────────────────────────────────────────────────────

def _side():
    return Side(border_style="thin", color="b0b8c1")

def _border():
    s = _side()
    return Border(left=s, right=s, top=s, bottom=s)

def _header_font():
    return Font(name="Tahoma", bold=True, color=WHITE, size=11)

def _data_font():
    return Font(name="Tahoma", size=10, color="1a1a2e")

def _arabic_align(wrap=False):
    return Alignment(horizontal="right", vertical="center",
                     text_rotation=0, wrap_text=wrap, readingOrder=2)

def _center_align():
    return Alignment(horizontal="center", vertical="center")

def _header_fill(hex_color: str = NAVY):
    return PatternFill("solid", fgColor=hex_color)

def _row_fill(row_idx: int):
    color = LIGHT_BLUE if row_idx % 2 == 0 else LIGHT_GREY
    return PatternFill("solid", fgColor=color)


# ─── Sheet-level helpers ──────────────────────────────────────────────────────

def set_rtl(ws: Worksheet):
    """Set sheet to right-to-left reading direction."""
    ws.sheet_view.rightToLeft = True


def freeze_header(ws: Worksheet):
    """Freeze the first row."""
    ws.freeze_panes = "A2"


def add_autofilter(ws: Worksheet):
    """Add Excel autofilter to the used range."""
    ws.auto_filter.ref = ws.dimensions


def set_tab_color(ws: Worksheet, hex_color: str):
    """Set the sheet tab colour."""
    ws.sheet_properties.tabColor = hex_color


# ─── Write DataFrame → Worksheet ─────────────────────────────────────────────
# FAST_THRESHOLD: rows above this use ws.append() (no per-cell styling = 10-20x faster)
_FAST_THRESHOLD = 3_000

def write_dataframe(
    ws: Worksheet,
    df: pl.DataFrame,
    header_color: str = NAVY,
    start_row: int = 1,
    start_col: int = 1,
    number_cols: Optional[List[str]] = None,
    highlight_col: Optional[str] = None,
    highlight_rules: Optional[dict] = None,
):
    """
    Write a Polars DataFrame into *ws* starting at (start_row, start_col).

    Performance strategy:
      • Pre-cache ALL style objects outside any loop (avoids millions of object creations).
      • Fast path for large DFs (>_FAST_THRESHOLD rows, start at row 1 col 1):
        uses ws.append() for raw values (10-20× faster), then applies alternating
        row color via ConditionalFormatting (rendered by Excel, not Python).
      • Styled path for small DFs (pivot tables, summaries): uses cached style objects.
    """
    number_cols = number_cols or []
    headers     = list(df.columns)
    num_col_set = set(number_cols)
    n_rows      = len(df)

    # ── Pre-create ALL style objects ONCE (huge speedup) ──────────────────
    hdr_font  = Font(name="Tahoma", bold=True, color=WHITE, size=11)
    hdr_fill  = PatternFill("solid", fgColor=header_color)
    hdr_brd   = _border()
    hdr_align = Alignment(horizontal="right", vertical="center", readingOrder=2)

    data_font  = Font(name="Tahoma", size=10, color="1a1a2e")
    data_align = Alignment(horizontal="right", vertical="center", readingOrder=2)
    fill_a     = PatternFill("solid", fgColor=LIGHT_BLUE)
    fill_b     = PatternFill("solid", fgColor=LIGHT_GREY)
    fills      = (fill_a, fill_b)

    # ── Write header row ───────────────────────────────────────────────────
    for col_idx, header in enumerate(headers, start=start_col):
        cell = ws.cell(row=start_row, column=col_idx, value=header)
        cell.font      = hdr_font
        cell.fill      = hdr_fill
        cell.border    = hdr_brd
        cell.alignment = hdr_align

    # ── Decide path ────────────────────────────────────────────────────────
    # Fast path only when data starts at top-left (append() must write after header)
    use_fast = n_rows > _FAST_THRESHOLD and start_row == 1 and start_col == 1

    if use_fast:
        # ── FAST PATH: raw append() — no per-cell styling ─────────────────
        for row_tuple in df.iter_rows():
            ws.append([("" if v is None else v) for v in row_tuple])

        # Apply Excel Table for alternating rows + borders + header styling.
        # Serializes faster than ConditionalFormatting rules for large data.
        end_col_letter = get_column_letter(start_col + len(headers) - 1)
        table_ref  = f"A{start_row}:{end_col_letter}{start_row + n_rows}"
        tbl_name   = f"Tbl{ws.title[:10].replace(' ','_')}"
        from openpyxl.worksheet.table import Table as XlTable, TableStyleInfo
        tbl = XlTable(displayName=tbl_name, ref=table_ref)
        tbl.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False, showLastColumn=False,
            showRowStripes=True, showColumnStripes=False,
        )
        ws.add_table(tbl)

        # Apply column-level number format for float/decimal columns (عدد العملاء etc.)
        # This is efficient — one assignment per column, not per cell
        for col_idx, (col_name, col_dtype) in enumerate(
            zip(df.columns, df.dtypes), start=start_col
        ):
            if col_name in num_col_set or col_dtype in (pl.Float32, pl.Float64):
                ws.column_dimensions[get_column_letter(col_idx)].number_format = '#,##0.000'


    else:
        # ── STYLED PATH: cached objects, no re-creation per cell ──────────
        for row_idx, row_tuple in enumerate(df.iter_rows(), start=start_row + 1):
            fill = fills[(row_idx - start_row) % 2]
            for col_idx, (col_name, value) in enumerate(zip(headers, row_tuple), start=start_col):
                cell = ws.cell(row=row_idx, column=col_idx, value="" if value is None else value)
                cell.font      = data_font
                cell.fill      = fill
                cell.alignment = data_align
                if col_name in num_col_set:
                    try:
                        cell.value = float(value) if value not in ("", None) else ""
                        cell.number_format = '#,##0.00'
                    except Exception:
                        pass

    # ── Auto-fit widths ────────────────────────────────────────────────────
    _auto_fit_columns(ws, df, start_row, start_col)




def _auto_fit_columns(
    ws: Worksheet,
    df: pl.DataFrame,
    start_row: int,
    start_col: int,
    min_width: int = 12,
    max_width: int = 45,
):
    for col_idx, col_name in enumerate(df.columns, start=start_col):
        col_letter = get_column_letter(col_idx)
        header_len = len(str(col_name))
        
        sample = df[col_name].head(200)
        max_data_len = sample.cast(pl.String).str.len_chars().max() if len(sample) else 0
        if max_data_len is None:
            max_data_len = 0
            
        width = min(max_width, max(min_width, header_len + 2, max_data_len + 2))
        ws.column_dimensions[col_letter].width = width


# ─── Conditional formatting ───────────────────────────────────────────────────

def highlight_errors(ws: Worksheet, col_letter: str, start_row: int, end_row: int):
    """Highlight non-empty cells in *col_letter* with red background."""
    ws.conditional_formatting.add(
        f"{col_letter}{start_row}:{col_letter}{end_row}",
        FormulaRule(
            formula=[f'LEN({col_letter}{start_row})>0'],
            fill=PatternFill("solid", fgColor="fce4e4"),
            font=Font(color="c0392b", bold=True, name="Tahoma"),
        ),
    )

def highlight_neglected(ws: Worksheet, col_letter: str, start_row: int, end_row: int):
    """Highlight cells equal to 'مهمل' with amber background."""
    ws.conditional_formatting.add(
        f"{col_letter}{start_row}:{col_letter}{end_row}",
        CellIsRule(
            operator="equal",
            formula=['"مهمل"'],
            fill=PatternFill("solid", fgColor="fff3cd"),
            font=Font(color="856404", bold=True, name="Tahoma"),
        ),
    )

def highlight_contact(ws: Worksheet, col_letter: str, start_row: int, end_row: int):
    """Green for contacted, red for not contacted."""
    rng = f"{col_letter}{start_row}:{col_letter}{end_row}"
    ws.conditional_formatting.add(rng, CellIsRule(
        operator="equal", formula=['"تم التوصل"'],
        fill=PatternFill("solid", fgColor="d4edda"),
        font=Font(color="155724", bold=True, name="Tahoma"),
    ))
    ws.conditional_formatting.add(rng, CellIsRule(
        operator="equal", formula=['"عدم التوصل"'],
        fill=PatternFill("solid", fgColor="f8d7da"),
        font=Font(color="721c24", bold=True, name="Tahoma"),
    ))

def colorscale_column(ws: Worksheet, col_letter: str, start_row: int, end_row: int):
    """Apply green-yellow-red colour scale to a numeric column."""
    ws.conditional_formatting.add(
        f"{col_letter}{start_row}:{col_letter}{end_row}",
        ColorScaleRule(
            start_type="min",  start_color="63be7b",
            mid_type="percentile", mid_value=50, mid_color="ffeb84",
            end_type="max",    end_color="f8696b",
        ),
    )


# ─── Chart factory ────────────────────────────────────────────────────────────

def create_bar_chart(
    ws: Worksheet,
    title: str,
    data_ref: Reference,
    cats_ref: Reference,
    anchor: str = "A1",
    width: int = 20,
    height: int = 12,
) -> BarChart:
    chart = BarChart()
    chart.type  = "col"
    chart.style = 10
    chart.title = title
    chart.y_axis.title = "العدد"
    chart.x_axis.numFmt = "@"
    chart.grouping = "clustered"
    chart.overlap  = 0

    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    chart.shape = 4
    chart.width  = width
    chart.height = height

    ws.add_chart(chart, anchor)
    return chart


def create_pie_chart(
    ws: Worksheet,
    title: str,
    data_ref: Reference,
    cats_ref: Reference,
    anchor: str = "A1",
    width: int = 18,
    height: int = 12,
) -> PieChart:
    chart = PieChart()
    chart.title  = title
    chart.style  = 10
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    chart.width  = width
    chart.height = height
    ws.add_chart(chart, anchor)
    return chart
