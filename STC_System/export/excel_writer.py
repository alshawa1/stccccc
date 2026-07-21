"""
export/excel_writer.py
──────────────────────
Orchestrates generation of the output Excel workbook.

Each public method writes one logical section.
The caller (gui/app.py) calls only the sections relevant to the selected task.

Sheet order when all tasks run:
  1. Dashboard
  2. Summary
  3. اخطاء النظام
  4. التوصل
  5. الإهمال
  6. الإضافة
  7. التسوية
  8. الحذف والتعديل
  9. الجدولة
  10. السحب والتدوير
  11. العملاء المستهدفة
"""
from __future__ import annotations

import logging
import os
from datetime import datetime
from typing import Any, Dict, Optional

import openpyxl
import polars as pl
from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from export.formatters import (
    AMBER, GREEN, NAVY, PURPLE, RED, TEAL,
    add_autofilter, colorscale_column, create_bar_chart, create_pie_chart,
    freeze_header, highlight_contact, highlight_errors, highlight_neglected,
    set_rtl, set_tab_color, write_dataframe,
)

_log = logging.getLogger("ExcelWriter")

# Tab colours per sheet
TAB_COLORS = {
    "Dashboard":        "1f6feb",
    "Summary":          "238636",
    "اخطاء النظام":    "da3633",
    "التوصل":           "28a745",
    "الإهمال":          "d29922",
    "الإضافة":          "8957e5",
    "التسوية":          "1abc9c",
    "الحذف والتعديل":  "e74c3c",
    "الجدولة":          "3498db",
    "السحب والتدوير":  "e67e22",
    "العملاء المستهدفة": "9b59b6",
}


class ExcelReportWriter:
    """
    Creates a single output workbook and exposes methods to populate
    each sheet independently. Call save() when done.
    """

    def __init__(self, output_path: str):
        self.output_path = output_path
        self.wb = Workbook()
        # Remove default blank sheet
        default = self.wb.active
        self.wb.remove(default)
        self._sheets_written: list[str] = []
        _log.info("🗂  تهيئة ملف الإخراج: %s", output_path)

    # ── Public: write methods ─────────────────────────────────────────────────

    # ── Dynamic KPI definitions per task ────────────────────────────────────────
    TASK_KPIS = {
        1: [  # أخطاء النظام
            ("👥 إجمالي العملاء",      "إجمالي العملاء",          "1f6feb"),
            ("❌ بأخطاء",              "عملاء بأخطاء",             "da3633"),
            ("✅ بدون أخطاء",          "عملاء بدون أخطاء",        "238636"),
            ("📊 نسبة الأخطاء",        "نسبة الأخطاء %",           "f0883e"),
            ("🆔 خطأ الرقم الرئيسي",  "خطأ الرقم الرئيسي",       "e74c3c"),
            ("⚠️ خطأ الحالة",          "خطأ الحالة",               "d29922"),
            ("📅 خطأ الوعد",           "خطأ الوعد",                "8957e5"),
            ("📝 خطأ الإفادة",         "خطأ الإفادة",              "3498db"),
        ],
        2: [  # التوصل وعدم التوصل
            ("👥 إجمالي العملاء",      "إجمالي العملاء",          "1f6feb"),
            ("✅ تم التوصل",           "تم التوصل",                "238636"),
            ("📵 عدم التوصل",          "عدم التوصل",              "da3633"),
            ("📊 نسبة التوصل %",       "نسبة التوصل %",            "1abc9c"),
        ],
        3: [  # الإهمال
            ("👥 إجمالي العملاء",      "إجمالي العملاء",          "1f6feb"),
            ("😴 مهمل",               "مهمل",                     "da3633"),
            ("✅ نشط",                "نشط",                      "238636"),
            ("📊 نسبة الإهمال %",      "نسبة الإهمال %",           "d29922"),
        ],
        4: [  # السدادات والتسوية
            ("➕ إضافة لمهارة",        "إضافة",                    "8957e5"),
            ("✅ مطابق",              "مطابق",                    "238636"),
            ("🔍 مراجعة",             "مراجعة",                   "f0883e"),
            ("🤝 تسوية كاملة",        "تسوية كاملة",              "1abc9c"),
            ("🔄 تسوية جزئية",        "تسوية جزئية",              "d29922"),
            ("✏️ تعديل",              "تعديل",                    "3498db"),
            ("🗑️ حذف",               "حذف",                      "da3633"),
            ("🟰 مطابق (تعديل)",      "مطابق (حذف/تعديل)",        "238636"),
        ],
        5: [  # الجدولة
            ("👥 إجمالي الجدولة",      "إجمالي العملاء",          "1f6feb"),
            ("📌 جدولة مؤكدة",        "جدولة مؤكدة",              "238636"),
            ("🔁 احتمال",             "احتمال",                   "d29922"),
            ("⚠️ متعثر",              "متعثر",                    "da3633"),
        ],
        6: [  # السحب والتدوير
            ("👥 إجمالي العملاء",      "إجمالي العملاء",          "1f6feb"),
            ("🔄 يُسحب",              "يُسحب",                    "e67e22"),
            ("🔁 يُدار",              "يُدار",                    "3498db"),
            ("✅ يبقى",               "يبقى",                     "238636"),
        ],
        7: [  # العملاء المستهدفة
            ("👥 إجمالي العملاء",      "إجمالي العملاء",          "1f6feb"),
            ("🌟 إيجابي",             "إيجابي",                   "238636"),
            ("🔴 سلبي",               "سلبي",                     "da3633"),
        ],
        8: [  # تشغيل جميع المهام
            ("👥 إجمالي العملاء",      "إجمالي العملاء",          "1f6feb"),
            ("❌ أخطاء النظام",        "عملاء بأخطاء",             "da3633"),
            ("😴 المهملون",            "مهمل",                     "d29922"),
            ("✅ تم التوصل",           "تم التوصل",                "238636"),
            ("📵 عدم التوصل",          "عدم التوصل",               "e74c3c"),
            ("➕ الإضافات",            "إضافة",                    "8957e5"),
            ("🤝 تسوية كاملة",        "تسوية كاملة",              "1abc9c"),
            ("📌 جدولة مؤكدة",        "جدولة مؤكدة",              "238636"),
            ("🌟 إيجابيون",            "إيجابي",                   "e91e8c"),
            ("🔄 يُسحب",              "يُسحب",                    "3498db"),
        ],
    }

    TASK_ARABIC_NAMES = {
        1: "أخطاء النظام",
        2: "التوصل وعدم التوصل",
        3: "الإهمال",
        4: "السدادات والتسوية",
        5: "الجدولة",
        6: "السحب والتدوير",
        7: "العملاء المستهدفة",
        8: "تقرير العمليات الكامل",
    }

    def write_dashboard(self, all_stats: Dict[str, Any], task_id: int = 8):
        """Write a dynamic Dashboard sheet tailored to the selected task."""
        ws = self._add_sheet("Dashboard", TAB_COLORS["Dashboard"])
        set_rtl(ws)

        task_name = self.TASK_ARABIC_NAMES.get(task_id, "العمليات")
        task_color = TAB_COLORS.get(list(TAB_COLORS.keys())[min(task_id, len(TAB_COLORS) - 1)], "1f6feb")

        # ── Title ──────────────────────────────────────────────────────────────
        ws.merge_cells("A1:L1")
        title = ws["A1"]
        title.value     = f"🏢  نظام أتمتة العمليات — مهارة × STC  |  {task_name}"
        title.font      = Font(name="Tahoma", size=15, bold=True, color="FFFFFF")
        title.fill      = PatternFill("solid", fgColor=NAVY)
        title.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 42

        # ── Subtitle ───────────────────────────────────────────────────────────
        ws.merge_cells("A2:L2")
        sub = ws["A2"]
        sub.value     = f"تاريخ التقرير: {datetime.now().strftime('%Y-%m-%d %H:%M')}  |  المهمة: {task_id} — {task_name}"
        sub.font      = Font(name="Tahoma", size=10, color="8b949e")
        sub.fill      = PatternFill("solid", fgColor="0d1117")
        sub.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[2].height = 22

        # ── KPI Cards ─────────────────────────────────────────────────────────
        kpi_definitions = self.TASK_KPIS.get(task_id, self.TASK_KPIS[8])
        kpi_data = [(label, all_stats.get(stat_key, 0), color)
                    for label, stat_key, color in kpi_definitions]

        row_start = 4
        cards_per_row = 5
        for i, (label, value, color) in enumerate(kpi_data):
            col = (i % cards_per_row) * 2 + 1
            r   = row_start + (i // cards_per_row) * 5

            # Top accent border
            ws.merge_cells(start_row=r, start_column=col, end_row=r, end_column=col + 1)
            top = ws.cell(row=r, column=col, value="")
            top.fill = PatternFill("solid", fgColor=color)
            ws.row_dimensions[r].height = 5

            # Label row
            ws.merge_cells(start_row=r+1, start_column=col, end_row=r+1, end_column=col+1)
            lc = ws.cell(row=r+1, column=col, value=label)
            lc.font      = Font(name="Tahoma", size=9, bold=True, color=color)
            lc.fill      = PatternFill("solid", fgColor="161b22")
            lc.alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[r+1].height = 18

            # Value row
            ws.merge_cells(start_row=r+2, start_column=col, end_row=r+3, end_column=col+1)
            vc = ws.cell(row=r+2, column=col, value=value)
            vc.font      = Font(name="Tahoma", size=26, bold=True, color=color)
            vc.fill      = PatternFill("solid", fgColor="161b22")
            vc.alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[r+2].height = 38
            ws.row_dimensions[r+3].height = 8

        # ── Full Stats Table ───────────────────────────────────────────────────
        # Find starting row for table (after KPI cards)
        rows_used = row_start + ((len(kpi_data) - 1) // cards_per_row + 1) * 5 + 2

        ws.merge_cells(f"A{rows_used}:L{rows_used}")
        hdr = ws.cell(row=rows_used, column=1, value="📊  تفاصيل جميع المؤشرات")
        hdr.font = Font(name="Tahoma", size=11, bold=True, color="c9d1d9")
        hdr.fill = PatternFill("solid", fgColor="21262d")
        hdr.alignment = Alignment(horizontal="right", vertical="center")
        ws.row_dimensions[rows_used].height = 22
        rows_used += 1

        col_colors = ["0d1117", "161b22"]
        for idx, (k, v) in enumerate(all_stats.items()):
            bg = col_colors[idx % 2]
            # Key (label) in columns A:F merged
            ws.merge_cells(f"A{rows_used}:F{rows_used}")
            kc = ws.cell(row=rows_used, column=1, value=str(k))
            kc.font      = Font(name="Tahoma", size=10, color="c9d1d9")
            kc.fill      = PatternFill("solid", fgColor=bg)
            kc.alignment = Alignment(horizontal="right", vertical="center")
            # Value in columns G:L merged
            ws.merge_cells(f"G{rows_used}:L{rows_used}")
            vc = ws.cell(row=rows_used, column=7, value=v)
            vc.font      = Font(name="Tahoma", size=10, bold=True, color="58a6ff")
            vc.fill      = PatternFill("solid", fgColor=bg)
            vc.alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[rows_used].height = 18
            rows_used += 1

        # ── Column widths ──────────────────────────────────────────────────────
        for c in range(1, 13):
            ws.column_dimensions[get_column_letter(c)].width = 17

        _log.info("✅ Dashboard sheet written (task %d — %s)", task_id, task_name)



    def write_summary(self, all_stats: Dict[str, Any]):
        """Write a flat summary table of all KPIs."""
        ws = self._add_sheet("Summary", TAB_COLORS["Summary"])
        set_rtl(ws)

        # Build summary rows — cast all values to str for uniform column type
        records = [{"المؤشر": str(k), "القيمة": str(v)} for k, v in all_stats.items()]
        if records:
            df_summary = pl.from_dicts(records)
        else:
            df_summary = pl.DataFrame({"المؤشر": pl.Series([], dtype=pl.String),
                                       "القيمة": pl.Series([], dtype=pl.String)})
        write_dataframe(ws, df_summary, header_color=GREEN)
        freeze_header(ws)
        add_autofilter(ws)
        _log.info("✅ Summary sheet written (%d KPIs)", len(all_stats))

    def write_errors(self, errors_data: pl.DataFrame):
        """اخطاء النظام sheet."""
        ws = self._add_sheet("اخطاء النظام", TAB_COLORS["اخطاء النظام"])
        set_rtl(ws)
        write_dataframe(ws, errors_data, header_color=RED)
        freeze_header(ws)
        add_autofilter(ws)

        # Highlight الخطأ column (new two-column output)
        if "الخطأ" in errors_data.columns:
            err_col_idx = list(errors_data.columns).index("الخطأ") + 1
            col_letter  = get_column_letter(err_col_idx)
            highlight_errors(ws, col_letter, 2, len(errors_data) + 1)

        # Highlight تصحيح الخطأ column with a soft green
        if "تصحيح الخطأ" in errors_data.columns:
            from openpyxl.formatting.rule import CellIsRule
            from openpyxl.styles import PatternFill, Font
            fix_col_idx = list(errors_data.columns).index("تصحيح الخطأ") + 1
            fix_letter  = get_column_letter(fix_col_idx)
            fix_range   = f"{fix_letter}2:{fix_letter}{len(errors_data) + 1}"
            ws.conditional_formatting.add(fix_range, CellIsRule(
                operator="notEqual", formula=['""'],
                fill=PatternFill("solid", fgColor="d4edda"),
                font=Font(color="155724", name="Tahoma", size=9),
            ))

        _log.info("✅ اخطاء النظام sheet written (%d rows)", len(errors_data))

    def write_contact(
        self,
        contact_data: pl.DataFrame,
        pivot_supervisor: pl.DataFrame,
        pivot_collector: pl.DataFrame,
        pivot_status: pl.DataFrame,
    ):
        """التوصل sheet + pivot sub-tables + charts."""
        ws = self._add_sheet("التوصل", TAB_COLORS["التوصل"])
        set_rtl(ws)
        write_dataframe(ws, contact_data, header_color=GREEN)
        freeze_header(ws)
        add_autofilter(ws)

        # Highlight contact status column
        if "حالة التوصل" in contact_data.columns:
            idx = list(contact_data.columns).index("حالة التوصل") + 1
            highlight_contact(ws, get_column_letter(idx), 2, len(contact_data) + 1)

        # Pivot — supervisor
        self._write_pivot_section(ws, pivot_supervisor, "حسب المشرف",
                                  start_row=len(contact_data) + 4, color=GREEN)
        # Pivot — collector
        row_after = len(contact_data) + 4 + len(pivot_supervisor) + 4
        self._write_pivot_section(ws, pivot_collector, "حسب المحصل",
                                  start_row=row_after, color=GREEN)

        _log.info("✅ التوصل sheet written")

    def write_neglect(
        self,
        neglect_only: pl.DataFrame,
        full_analysis: pl.DataFrame,
        pivot_summary: pl.DataFrame,
        pivot_supervisor: pl.DataFrame,
        pivot_collector: pl.DataFrame,
        pivot_status: pl.DataFrame,
        pivot_branch: pl.DataFrame,
        pivot_portfolio: pl.DataFrame,
        pivot_days: pl.DataFrame,
    ):
        """الإهمال sheets: 'الإهمال' (neglected only) and 'تحليل الإهمال' (both with pivots/charts)."""
        # --- Sheet 1: الإهمال (Only neglected) ---
        ws_only = self._add_sheet("الإهمال", TAB_COLORS["الإهمال"])
        set_rtl(ws_only)
        write_dataframe(ws_only, neglect_only, header_color=AMBER)
        freeze_header(ws_only)
        add_autofilter(ws_only)

        # Highlight neglect status in only sheet
        status_col_name = "حالة الاهمال" if "حالة الاهمال" in neglect_only.columns else "حالة الإهمال"
        if status_col_name in neglect_only.columns:
            idx_neg = list(neglect_only.columns).index(status_col_name) + 1
            highlight_neglected(ws_only, get_column_letter(idx_neg), 2, len(neglect_only) + 1)

        if "عدد أيام الإهمال" in neglect_only.columns:
            idx_days = list(neglect_only.columns).index("عدد أيام الإهمال") + 1
            colorscale_column(ws_only, get_column_letter(idx_days), 2, len(neglect_only) + 1)

        # --- Sheet 2: تحليل الإهمال (Both neglected and active) ---
        ws_anal = self._add_sheet("تحليل الإهمال", "f39c12")
        set_rtl(ws_anal)
        write_dataframe(ws_anal, full_analysis, header_color=AMBER)
        freeze_header(ws_anal)
        add_autofilter(ws_anal)

        # Highlight neglect status in analysis sheet
        status_col_name_anal = "حالة الاهمال" if "حالة الاهمال" in full_analysis.columns else "حالة الإهمال"
        if status_col_name_anal in full_analysis.columns:
            idx_neg = list(full_analysis.columns).index(status_col_name_anal) + 1
            highlight_neglected(ws_anal, get_column_letter(idx_neg), 2, len(full_analysis) + 1)

        if "عدد أيام الإهمال" in full_analysis.columns:
            idx_days = list(full_analysis.columns).index("عدد أيام الإهمال") + 1
            colorscale_column(ws_anal, get_column_letter(idx_days), 2, len(full_analysis) + 1)

        # Write pivots below data in ws_anal
        current_row = len(full_analysis) + 4

        # 1. Summary
        row_summary = current_row
        self._write_pivot_section(ws_anal, pivot_summary, "ملخص الإهمال", row_summary, color=AMBER, start_col=1)
        current_row += len(pivot_summary) + 4

        # 2. Supervisor
        row_supervisor = current_row
        self._write_pivot_section(ws_anal, pivot_supervisor, "حسب المشرف", row_supervisor, color=AMBER, start_col=1)
        current_row += len(pivot_supervisor) + 4

        # 3. Collector
        row_collector = current_row
        self._write_pivot_section(ws_anal, pivot_collector, "حسب المحصل", row_collector, color="34495e", start_col=1)
        current_row += len(pivot_collector) + 4

        # 4. Main Status
        row_status = current_row
        self._write_pivot_section(ws_anal, pivot_status, "حسب الحالة الرئيسية", row_status, color="7f8c8d", start_col=1)
        current_row += len(pivot_status) + 4

        # 5. Branch
        row_branch = current_row
        self._write_pivot_section(ws_anal, pivot_branch, "حسب الفرع", row_branch, color="1abc9c", start_col=1)
        current_row += len(pivot_branch) + 4

        # 6. Portfolio
        row_portfolio = current_row
        self._write_pivot_section(ws_anal, pivot_portfolio, "حسب المحفظة", row_portfolio, color="9b59b6", start_col=1)
        current_row += len(pivot_portfolio) + 4

        # 7. Days distribution
        row_days = current_row
        self._write_pivot_section(ws_anal, pivot_days, "توزيع أيام الإهمال", row_days, color="d35400", start_col=1)
        current_row += len(pivot_days) + 4

        # Add charts side-by-side or stacked at Column I on ws_anal
        chart_start_row = len(full_analysis) + 4

        # Chart 1: Pie Chart (مهمل / غير مهمل)
        if len(pivot_summary) > 0:
            c1_data = Reference(ws_anal, min_col=2, min_row=row_summary + 1, max_row=row_summary + 1 + len(pivot_summary))
            c1_cats = Reference(ws_anal, min_col=1, min_row=row_summary + 2, max_row=row_summary + 1 + len(pivot_summary))
            create_pie_chart(ws_anal, "نسبة الإهمال (Pie)", c1_data, c1_cats, anchor=f"I{chart_start_row}", width=11, height=7)

        # Chart 2: Bar Chart (عدد المهملين لكل مشرف)
        if len(pivot_supervisor) > 0:
            c2_data = Reference(ws_anal, min_col=2, min_row=row_supervisor + 1, max_row=row_supervisor + 1 + len(pivot_supervisor))
            c2_cats = Reference(ws_anal, min_col=1, min_row=row_supervisor + 2, max_row=row_supervisor + 1 + len(pivot_supervisor))
            create_bar_chart(ws_anal, "عدد المهملين لكل مشرف", c2_data, c2_cats, anchor=f"I{chart_start_row + 9}", width=11, height=7)

        # Chart 3: Bar Chart (عدد المهملين لكل محصل)
        if len(pivot_collector) > 0:
            c3_data = Reference(ws_anal, min_col=2, min_row=row_collector + 1, max_row=row_collector + 1 + len(pivot_collector))
            c3_cats = Reference(ws_anal, min_col=1, min_row=row_collector + 2, max_row=row_collector + 1 + len(pivot_collector))
            create_bar_chart(ws_anal, "عدد المهملين لكل محصل", c3_data, c3_cats, anchor=f"I{chart_start_row + 18}", width=11, height=7)

        # Chart 4: Line Chart (توزيع أيام الإهمال)
        if len(pivot_days) > 0:
            from openpyxl.chart import LineChart
            chart = LineChart()
            chart.title = "توزيع أيام الإهمال"
            chart.y_axis.title = "عدد العملاء"
            chart.x_axis.title = "أيام الإهمال"
            c4_data = Reference(ws_anal, min_col=2, min_row=row_days + 1, max_row=row_days + 1 + len(pivot_days))
            c4_cats = Reference(ws_anal, min_col=1, min_row=row_days + 2, max_row=row_days + 1 + len(pivot_days))
            chart.add_data(c4_data, titles_from_data=True)
            chart.set_categories(c4_cats)
            chart.width = 11
            chart.height = 7
            ws_anal.add_chart(chart, f"I{chart_start_row + 27}")

        _log.info("✅ الإهمال sheets written")

    def write_addition(self, addition_data: pl.DataFrame):
        """الإضافة sheet."""
        ws = self._add_sheet("الإضافة", TAB_COLORS["الإضافة"])
        set_rtl(ws)
        number_cols = ["الشركة", "مهارة", "الفرق", "مبلغ المديونية"]
        write_dataframe(ws, addition_data, header_color=PURPLE,
                        number_cols=[c for c in number_cols if c in addition_data.columns])
        freeze_header(ws)
        add_autofilter(ws)
        _log.info("✅ الإضافة sheet written (%d rows)", len(addition_data))

    def write_settlement(self, settlement_data: pl.DataFrame):
        """التسوية sheet."""
        ws = self._add_sheet("التسوية", TAB_COLORS["التسوية"])
        set_rtl(ws)
        write_dataframe(ws, settlement_data, header_color=TEAL)
        freeze_header(ws)
        add_autofilter(ws)

        if "نوع التسوية" in settlement_data.columns:
            idx = list(settlement_data.columns).index("نوع التسوية") + 1
            col_letter = get_column_letter(idx)
            from openpyxl.formatting.rule import CellIsRule
            from openpyxl.styles import PatternFill, Font
            rng = f"{col_letter}2:{col_letter}{len(settlement_data) + 1}"
            ws.conditional_formatting.add(rng, CellIsRule(
                operator="equal", formula=['"تسوية كاملة"'],
                fill=PatternFill("solid", fgColor="d4edda"),
                font=Font(color="155724", bold=True, name="Tahoma"),
            ))
            ws.conditional_formatting.add(rng, CellIsRule(
                operator="equal", formula=['"تسوية جزئية"'],
                fill=PatternFill("solid", fgColor="fff3cd"),
                font=Font(color="856404", bold=True, name="Tahoma"),
            ))

        _log.info("✅ التسوية sheet written (%d rows)", len(settlement_data))

    def write_edit_delete(self, edit_delete_data: pl.DataFrame):
        """الحذف والتعديل sheet."""
        ws = self._add_sheet("الحذف والتعديل", TAB_COLORS["الحذف والتعديل"])
        set_rtl(ws)
        number_cols = ["مهارة", "الشركة", "الفرق", "مبلغ التعديل"]
        write_dataframe(ws, edit_delete_data, header_color=RED,
                        number_cols=[c for c in number_cols if c in edit_delete_data.columns])
        freeze_header(ws)
        add_autofilter(ws)
        _log.info("✅ الحذف والتعديل sheet written (%d rows)", len(edit_delete_data))

    def write_scheduling(
        self,
        scheduling_data: pl.DataFrame,
        pivot_type: pl.DataFrame,
        pivot_month: pl.DataFrame,
        pivot_year: pl.DataFrame,
        pivot_count: pl.DataFrame,
        pivot_supervisor: pl.DataFrame,
    ):
        """الجدولة sheet with 5 pivots and 4 charts calculated dynamically to prevent overlaps."""
        ws = self._add_sheet("الجدولة", TAB_COLORS["الجدولة"])
        set_rtl(ws)
        write_dataframe(ws, scheduling_data, header_color="3498db")
        freeze_header(ws)
        add_autofilter(ws)

        start_row = len(scheduling_data) + 4

        # 1. Write Pivots in a Grid below data table
        # Pivot 1: Type (Col A)
        self._write_pivot_section(ws, pivot_type, "نوع الجدولة", start_row, color="3498db", start_col=1)
        # Pivot 2: Month (Col E)
        self._write_pivot_section(ws, pivot_month, "آخر دفعة حسب الشهر", start_row, color="1abc9c", start_col=5)
        # Pivot 3: Year (Col H)
        self._write_pivot_section(ws, pivot_year, "آخر دفعة حسب السنة", start_row, color="9b59b6", start_col=8)
        # Pivot 4: Count (Col K)
        self._write_pivot_section(ws, pivot_count, "حسب عدد الدفعات", start_row, color="e67e22", start_col=11)

        # Pivot 5: Supervisor (row is dynamic based on maximum length of grid pivots)
        max_len = max(len(pivot_type), len(pivot_month), len(pivot_year), len(pivot_count))
        row_supervisor = start_row + max_len + 4
        self._write_pivot_section(ws, pivot_supervisor, "توزيع الفئات حسب المشرف", row_supervisor, color="34495e", start_col=1)

        # 2. Add Charts
        # Chart 1: Pie Chart (نوع الجدولة) -> anchored at row_supervisor + len(pivot_supervisor) + 4, Col A
        chart_start_row = row_supervisor + len(pivot_supervisor) + 4
        if len(pivot_type) > 0:
            c1_data = Reference(ws, min_col=2, min_row=start_row + 1, max_row=start_row + 1 + len(pivot_type))
            c1_cats = Reference(ws, min_col=1, min_row=start_row + 2, max_row=start_row + 1 + len(pivot_type))
            create_pie_chart(ws, "توزيع فئات الجدولة (Pie)", c1_data, c1_cats, anchor=f"A{chart_start_row}", width=11, height=7)

        # Chart 2: Column Chart (نوع الجدولة) -> anchored at Col E
        if len(pivot_type) > 0:
            c2_data = Reference(ws, min_col=2, min_row=start_row + 1, max_row=start_row + 1 + len(pivot_type))
            c2_cats = Reference(ws, min_col=1, min_row=start_row + 2, max_row=start_row + 1 + len(pivot_type))
            create_bar_chart(ws, "عدد العملاء حسب نوع الجدولة", c2_data, c2_cats, anchor=f"E{chart_start_row}", width=11, height=7)

        # Chart 3: Monthly Distribution -> anchored at chart_start_row + 8, Col A
        if len(pivot_month) > 0:
            c3_data = Reference(ws, min_col=6, min_row=start_row + 1, max_row=start_row + 1 + len(pivot_month))
            c3_cats = Reference(ws, min_col=5, min_row=start_row + 2, max_row=start_row + 1 + len(pivot_month))
            create_bar_chart(ws, "توزيع العملاء شهرياً", c3_data, c3_cats, anchor=f"A{chart_start_row + 8}", width=11, height=7)

        # Chart 4: Payment Frequency -> anchored at chart_start_row + 8, Col E
        if len(pivot_count) > 0:
            c4_data = Reference(ws, min_col=12, min_row=start_row + 1, max_row=start_row + 1 + len(pivot_count))
            c4_cats = Reference(ws, min_col=11, min_row=start_row + 2, max_row=start_row + 1 + len(pivot_count))
            create_bar_chart(ws, "تكرار الدفعات للعملاء", c4_data, c4_cats, anchor=f"E{chart_start_row + 8}", width=11, height=7)

        _log.info("✅ الجدولة sheet written")

    def write_withdrawal(
        self,
        withdrawal_data: pl.DataFrame,
        pivot_supervisor: pl.DataFrame,
    ):
        """السحب والتدوير sheet."""
        ws = self._add_sheet("السحب والتدوير", TAB_COLORS["السحب والتدوير"])
        set_rtl(ws)
        write_dataframe(ws, withdrawal_data, header_color="e67e22")
        freeze_header(ws)
        add_autofilter(ws)
        self._write_pivot_section(ws, pivot_supervisor, "حسب المشرف",
                                  start_row=len(withdrawal_data) + 4, color="e67e22")
        _log.info("✅ السحب والتدوير sheet written (%d rows)", len(withdrawal_data))

    def write_targets(
        self,
        target_data: pl.DataFrame,
        pivot_supervisor: pl.DataFrame,
    ):
        """العملاء المستهدفة sheet."""
        ws = self._add_sheet("العملاء المستهدفة", TAB_COLORS["العملاء المستهدفة"])
        set_rtl(ws)
        write_dataframe(ws, target_data, header_color="9b59b6")
        freeze_header(ws)
        add_autofilter(ws)
        self._write_pivot_section(ws, pivot_supervisor, "حسب المشرف",
                                  start_row=len(target_data) + 4, color="9b59b6")
        _log.info("✅ العملاء المستهدفة sheet written (%d rows)", len(target_data))

    # ── Save ──────────────────────────────────────────────────────────────────

    def save(self):
        """Save the workbook to disk."""
        self.wb.save(self.output_path)
        _log.info("💾 Workbook saved → %s", self.output_path)

    # ── Private helpers ───────────────────────────────────────────────────────

    def _add_sheet(self, name: str, tab_color: str = NAVY) -> object:
        ws = self.wb.create_sheet(title=name)
        set_tab_color(ws, tab_color)
        self._sheets_written.append(name)
        return ws

    def _write_pivot_section(
        self,
        ws,
        pivot_df: pl.DataFrame,
        section_title: str,
        start_row: int,
        color: str = NAVY,
        start_col: int = 1,
    ):
        if pivot_df is None or len(pivot_df) == 0 or len(pivot_df.columns) == 0:
            return

        # Section title
        ws.merge_cells(
            start_row=start_row, start_column=start_col,
            end_row=start_row,   end_column=start_col + min(len(pivot_df.columns), 8) - 1,
        )
        title_cell = ws.cell(row=start_row, column=start_col, value=f"📊 {section_title}")
        title_cell.font      = Font(name="Tahoma", size=12, bold=True, color="FFFFFF")
        title_cell.fill      = PatternFill("solid", fgColor=color)
        title_cell.alignment = Alignment(horizontal="right", vertical="center",
                                         readingOrder=2)
        ws.row_dimensions[start_row].height = 25

        write_dataframe(ws, pivot_df, header_color=color, start_row=start_row + 1, start_col=start_col)
