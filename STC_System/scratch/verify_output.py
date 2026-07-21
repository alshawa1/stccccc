"""Verify the output Excel file has correct sheets and data."""
import sys, os
sys.stdout.reconfigure(encoding='utf-8')

import openpyxl

path = r'c:\Users\dell\Downloads\فايلات مهاره\test_xlsxwriter_neglect.xlsx'
wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

print(f"File: {os.path.basename(path)}")
print(f"Sheets: {wb.sheetnames}")
print()

for sname in wb.sheetnames:
    ws = wb[sname]
    # count used rows
    max_row = ws.max_row
    max_col = ws.max_column
    # read first row (headers)
    headers = []
    first_data = []
    for row in ws.iter_rows(min_row=1, max_row=2, values_only=True):
        if not headers:
            headers = [str(c) if c is not None else '' for c in row]
        else:
            first_data = [str(c)[:30] if c is not None else '' for c in row]
    
    print(f"=== {sname} ===")
    print(f"  Rows: {max_row}, Cols: {max_col}")
    if "عدد العملاء" in headers:
        idx = headers.index("عدد العملاء")
        print(f"  ✅ عدد العملاء column at index {idx}: value={first_data[idx] if first_data else 'N/A'}")
    else:
        print(f"  ℹ️  لا يوجد عمود عدد العملاء في هذا الشيت")
    if headers:
        print(f"  First 5 headers: {headers[:5]}")
    print()

wb.close()
print("✅ Verification complete!")
